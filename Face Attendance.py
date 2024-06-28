import face_recognition
import cv2
import numpy as np
import csv
from datetime import datetime
import openpyxl

garbhit_image = face_recognition.load_image_file("C:/Users/kaush/Downloads/Photos/Garbhit.jpg")
garbhit_encoding = face_recognition.face_encodings(garbhit_image)[0]

shivank_image = face_recognition.load_image_file("C:/Users/kaush/Downloads/Photos/Shivank.jpg")
shivank_encoding = face_recognition.face_encodings(shivank_image)[0]

mohit_image = face_recognition.load_image_file("C:/Users/kaush/Downloads/Photos/Mohit.jpg")
mohit_encoding = face_recognition.face_encodings(mohit_image)[0]

mukul_image = face_recognition.load_image_file("C:/Users/kaush/Downloads/Photos/Mukul.jpg")
mukul_encoding = face_recognition.face_encodings(mukul_image)[0]


video_capture = cv2.VideoCapture(0)

known_face_encodings = [
    garbhit_encoding,
    shivank_encoding,
    mohit_encoding,
    mukul_encoding
]

known_face_names = [
    "Garbhit",
    "Shivank",
    "Mohit" ,
    "Mukul"
]

students = known_face_names.copy()

now = datetime.now()
current_date = now.strftime("%Y-%m-%d")
excel_filename = current_date + '.xlsx'
workbook = openpyxl.Workbook()
sheet = workbook.active

sheet['A1'] = 'Name'
sheet['B1'] = 'Date'
sheet['C1'] = 'Time'
sheet['D1'] = 'Status'

while True:
    _, frame = video_capture.read()
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    for face_encoding in face_encodings:

        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]

        face_names.append(name)

        if name in students:
            font = cv2.FONT_HERSHEY_SIMPLEX
            bottomLeftCornerOfText = (10, 100)
            fontScale = 1.5
            fontColor = (0, 255, 0)  
            thickness = 3
            lineType = 2

            cv2.putText(frame, name + ' Present',
                        bottomLeftCornerOfText,
                        font,
                        fontScale,
                        fontColor,
                        thickness,
                        lineType)

            students.remove(name)
            current_time = now.strftime("%H-%M-%S")

          
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=name)
            sheet.cell(row=next_row, column=2, value=current_date)
            sheet.cell(row=next_row, column=3, value=current_time)
            sheet.cell(row=next_row, column=4, value='Present')


    workbook.save(excel_filename)

    cv2.imshow('Show Faces', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
